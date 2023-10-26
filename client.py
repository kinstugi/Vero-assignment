import requests
import argparse
import openpyxl
from openpyxl.styles import PatternFill
import datetime
from custom_utils import months_passed

parser = argparse.ArgumentParser()
parser.add_argument('-c', '--colored', action='store_true', help='')
parser.add_argument('-k', '--keys', nargs='*', help='')

args = parser.parse_args()
url = "http://127.0.0.1:5000"
print(parser)

def get_fields(item: dict)->dict:
    res = {'rnr': item['rnr']}
    keys = args.keys if args.keys else []
    for k in keys:
        res[k] = item.get(k, None)
    return res

if __name__ == "__main__":
    vh_file = open('./files/vehicles.csv', 'rb')
    res = requests.post(url, files={'file': vh_file})
    
    if 200 <= res.status_code <= 299:
        data = res.json()
        data.sort(key = lambda x: x['gruppe'])
        new_data = list(map(get_fields, data))

        workbook = openpyxl.Workbook()
        worksheet = workbook.active

        for cc, (k,v) in enumerate(new_data[0].items()):
            worksheet.cell(row=1, column= cc+ 1).value = k
        
        for rr,item in enumerate(new_data):
            for cc, (k,v) in enumerate(item.items()): 
                worksheet.cell(row=rr+2, column=cc + 1).value = v
            
            fill = None
            if args.colored and data[rr]["hu"]:
                mm = months_passed(data[rr]["hu"])
                if mm <= 3:
                    fill = PatternFill(start_color='007500', end_color='007500', fill_type='solid')
                elif mm <= 12:
                    fill = PatternFill(start_color='FFA500', end_color='FFA500', fill_type= 'solid')
                else:
                    fill = PatternFill(start_color='b30000', end_color='b30000', fill_type='solid')
            
            if fill:
                for cell in worksheet[rr+2]:    
                    cell.fill = fill
                

        dt_now = datetime.datetime.now().isoformat()
        workbook.save(f'./output/vehicles_{dt_now}.xlsx')